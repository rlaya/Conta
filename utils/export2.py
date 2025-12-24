# utils/export2.py
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

def exportar_asientos_a_excel(asientos):
    wb = Workbook()
    ws = wb.active
    ws.title = "Asientos Contables"

    # Encabezados
    headers = ["ID", "Fecha", "Diario", "Concepto", "Referencia", "Creador", "Cliente", "Proveedor"]
    ws.append(headers)
    
    # Estilo
    for col in range(1, len(headers) + 1):
        ws.cell(1, col).font = Font(bold=True)
        ws.cell(1, col).alignment = Alignment(horizontal="center")

    # Datos
    for a in asientos:
        ws.append([
            a[0],  # id_asiento
            a[1].strftime('%Y-%m-%d'),  # fecha
            a[4],  # diario
            a[2],  # concepto
            a[3] or '',  # referencia
            a[5],  # creador
            a[6] or '',  # cliente
            a[7] or ''   # proveedor
        ])

    # Ajustar ancho
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = min(adjusted_width, 30)

    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output